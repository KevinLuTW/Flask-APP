from flask import Flask, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)

# 创建一个 Excel 文件来保存联系方式
def create_or_update_excel_file(email, phone):
    excel_file = "contacts.xlsx"
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Email", "Phone", "Download Time"])

    ws.append([email, phone, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(excel_file)

# 定义一个简单的表单页面，用户输入联系方式
@app.route('/')
def index():
    return render_template('index.html')

# 处理表单提交
@app.route('/download', methods=['POST'])
def download():
    email = request.form['email']
    phone = request.form['phone']
    
    # 保存联系方式到 Excel 文件中
    create_or_update_excel_file(email, phone)
    
    # 返回要下载的文件
    return send_file('your_file.rar', as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
